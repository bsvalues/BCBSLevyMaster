"""
Levy Export Parser for the SaaS Levy Calculation System.

This module provides classes for parsing levy export files in various formats,
including text (.txt), Excel (.xls, .xlsx), and XML formats.
"""

import os
import re
import csv
import logging
import tempfile
from enum import Enum
from typing import List, Dict, Any, Union, Optional, Tuple
from datetime import datetime
from pathlib import Path

# Third-party libraries
import openpyxl
import pandas as pd
import xlrd

# Setup logging
logger = logging.getLogger(__name__)


class LevyExportFormat(Enum):
    """Enumeration of supported levy export file formats."""
    TXT = 'txt'
    XLS = 'xls'
    XLSX = 'xlsx'
    XML = 'xml'
    UNKNOWN = 'unknown'


class LevyRecord:
    """A single levy record extracted from a levy export file."""
    
    def __init__(self, data: Dict[str, Any]):
        """
        Initialize a levy record with the provided data.
        
        Args:
            data: Dictionary of levy record data
        """
        self.data = data
        
    def __getitem__(self, key: str) -> Any:
        """
        Get a value from the record data.
        
        Args:
            key: The key to look up
            
        Returns:
            The value for the key or None if not found
        """
        return self.data.get(key)
    
    def get(self, key: str, default: Any = None) -> Any:
        """
        Get a value from the record data with a default.
        
        Args:
            key: The key to look up
            default: Default value if key not found
            
        Returns:
            The value for the key or the default if not found
        """
        return self.data.get(key, default)


class LevyExportData:
    """Container for levy export data extracted from a file."""
    
    def __init__(self, records: List[Dict[str, Any]], metadata: Optional[Dict[str, Any]] = None):
        """
        Initialize with records and optional metadata.
        
        Args:
            records: List of levy record dictionaries
            metadata: Optional metadata about the export
        """
        self.records = [LevyRecord(record) for record in records]
        self.metadata = metadata or {}
        
    def __len__(self) -> int:
        """Get the number of records."""
        return len(self.records)
    
    def get_years(self) -> List[int]:
        """
        Get a list of all years in the data.
        
        Returns:
            List of years found in the records
        """
        years = set()
        for record in self.records:
            year = record['year']
            if year:
                try:
                    years.add(int(year))
                except (ValueError, TypeError):
                    pass
        return sorted(list(years))
    
    def get_tax_districts(self) -> List[str]:
        """
        Get a list of all tax districts in the data.
        
        Returns:
            List of unique tax district IDs
        """
        districts = set()
        for record in self.records:
            district = record['tax_district_id']
            if district:
                districts.add(str(district))
        return sorted(list(districts))
    
    def get_levy_codes(self) -> List[str]:
        """
        Get a list of all levy codes in the data.
        
        Returns:
            List of unique levy codes
        """
        codes = set()
        for record in self.records:
            code = record['levy_cd']
            if code:
                codes.add(str(code))
        return sorted(list(codes))


class LevyExportParser:
    """Parser for levy export files in various formats."""
    
    @classmethod
    def detect_format(cls, file_path: Union[str, Path]) -> LevyExportFormat:
        """
        Detect the format of a levy export file.
        
        Args:
            file_path: Path to the file
            
        Returns:
            The detected file format as a LevyExportFormat enum
        """
        file_path = Path(file_path)
        extension = file_path.suffix.lower().lstrip('.')
        
        if extension == 'txt':
            return LevyExportFormat.TXT
        elif extension == 'xls':
            return LevyExportFormat.XLS
        elif extension == 'xlsx':
            return LevyExportFormat.XLSX
        elif extension == 'xml':
            return LevyExportFormat.XML
        else:
            return LevyExportFormat.UNKNOWN
    
    @classmethod
    def parse_file(cls, file_path: Union[str, Path]) -> LevyExportData:
        """
        Parse a levy export file.
        
        Args:
            file_path: Path to the file
            
        Returns:
            LevyExportData object containing the parsed data
            
        Raises:
            ValueError: If the file format is not supported
        """
        file_path = Path(file_path)
        format = cls.detect_format(file_path)
        
        if format == LevyExportFormat.TXT:
            return cls._parse_txt(file_path)
        elif format == LevyExportFormat.XLS:
            return cls._parse_xls(file_path)
        elif format == LevyExportFormat.XLSX:
            return cls._parse_xlsx(file_path)
        elif format == LevyExportFormat.XML:
            return cls._parse_xml(file_path)
        else:
            raise ValueError(f"Unsupported file format: {format}")
    
    @classmethod
    def _parse_txt(cls, file_path: Path) -> LevyExportData:
        """
        Parse a text format levy export file.
        
        Args:
            file_path: Path to the TXT file
            
        Returns:
            LevyExportData object containing the parsed data
        """
        logger.info(f"Parsing TXT levy export file: {file_path}")
        
        records = []
        year = datetime.now().year
        header_pattern = re.compile(r'^\s*LEVY CODE\s+RATE\s+LEVY\s+VALUE\s*$', re.IGNORECASE)
        
        try:
            with open(file_path, 'r', encoding='utf-8-sig') as f:
                lines = f.readlines()
                
                # Scan for headers and extract year if available
                for i, line in enumerate(lines[:20]):  # Check first 20 lines for headers
                    if re.search(r'(LEVY|TAX)\s+YEAR\s*[:=\s]\s*(\d{4})', line, re.IGNORECASE):
                        year_match = re.search(r'(\d{4})', line)
                        if year_match:
                            year = int(year_match.group(1))
                            logger.info(f"Found year in header: {year}")
                    
                    if header_pattern.search(line):
                        # Found the column headers, start parsing from next line
                        data_lines = lines[i+1:]
                        break
                else:
                    # If no header found, assume all lines are data
                    data_lines = lines
                
                # Parse data lines
                for line in data_lines:
                    line = line.strip()
                    if not line or line.startswith('#'):
                        continue
                    
                    # Try to parse fixed-width format
                    match = re.match(r'^\s*(\S+)\s+(\d*\.?\d*)\s+(\d*\.?\d*)\s+(\d*\.?\d*)', line)
                    if match:
                        levy_cd, rate, levy, value = match.groups()
                        
                        # Clean up and convert values
                        try:
                            rate = float(rate) if rate else None
                            levy = float(levy) if levy else None
                            value = float(value) if value else None
                        except ValueError:
                            logger.warning(f"Failed to parse values from line: {line}")
                            continue
                        
                        # Check for linked levy code
                        levy_cd_linked = None
                        if '/' in levy_cd:
                            levy_cd, levy_cd_linked = levy_cd.split('/', 1)
                        
                        record = {
                            'tax_district_id': levy_cd,
                            'levy_cd': levy_cd,
                            'levy_cd_linked': levy_cd_linked,
                            'levy_rate': rate,
                            'levy_amount': levy,
                            'assessed_value': value,
                            'year': year,
                            'source': 'txt'
                        }
                        records.append(record)
        
        except Exception as e:
            logger.error(f"Error parsing TXT file {file_path}: {str(e)}")
            raise
        
        logger.info(f"Parsed {len(records)} records from TXT file")
        return LevyExportData(records, {'format': 'txt', 'year': year})
    
    @classmethod
    def _parse_xls(cls, file_path: Path) -> LevyExportData:
        """
        Parse an Excel .xls format levy export file.
        
        Args:
            file_path: Path to the XLS file
            
        Returns:
            LevyExportData object containing the parsed data
        """
        logger.info(f"Parsing XLS levy export file: {file_path}")
        
        try:
            # Open the workbook and select the first sheet
            wb = xlrd.open_workbook(file_path)
            sheet = wb.sheet_by_index(0)
            
            # Attempt to find header row
            header_row = None
            year = datetime.now().year
            
            for row_idx in range(min(20, sheet.nrows)):  # Check first 20 rows
                row_values = [str(cell).strip().upper() for cell in sheet.row_values(row_idx)]
                row_text = ' '.join(row_values)
                
                # Look for year in header rows
                if re.search(r'(LEVY|TAX)\s+YEAR\s*[:=\s]\s*(\d{4})', row_text, re.IGNORECASE):
                    year_match = re.search(r'(\d{4})', row_text)
                    if year_match:
                        year = int(year_match.group(1))
                        logger.info(f"Found year in header: {year}")
                
                # Look for column headers
                if 'LEVY CODE' in row_text and 'RATE' in row_text:
                    header_row = row_idx
                    logger.info(f"Found header row at index {header_row}")
                    break
            
            # If header not found, try to infer it
            if header_row is None:
                logger.warning("No header row found, attempting to infer column structure")
                header_row = 0
            
            # Get column indices
            header_values = [str(cell).strip().upper() for cell in sheet.row_values(header_row)]
            
            col_indices = {
                'levy_cd': next((i for i, h in enumerate(header_values) if 'LEVY CODE' in h or 'CODE' in h), 0),
                'rate': next((i for i, h in enumerate(header_values) if 'RATE' in h), 1),
                'levy': next((i for i, h in enumerate(header_values) if 'LEVY' in h and 'CODE' not in h), 2),
                'value': next((i for i, h in enumerate(header_values) if 'VALUE' in h or 'ASSESSED' in h), 3)
            }
            
            # Parse data rows
            records = []
            for row_idx in range(header_row + 1, sheet.nrows):
                row_values = sheet.row_values(row_idx)
                
                if not row_values[col_indices['levy_cd']]:  # Skip empty rows
                    continue
                
                # Get values from appropriate columns
                levy_cd = str(row_values[col_indices['levy_cd']]).strip()
                
                # Handle potential float formatting issues
                try:
                    rate = float(row_values[col_indices['rate']]) if row_values[col_indices['rate']] else None
                except (ValueError, TypeError):
                    rate = None
                
                try:
                    levy = float(row_values[col_indices['levy']]) if row_values[col_indices['levy']] else None
                except (ValueError, TypeError):
                    levy = None
                
                try:
                    value = float(row_values[col_indices['value']]) if row_values[col_indices['value']] else None
                except (ValueError, TypeError):
                    value = None
                
                # Check for linked levy code
                levy_cd_linked = None
                if '/' in levy_cd:
                    levy_cd, levy_cd_linked = levy_cd.split('/', 1)
                
                record = {
                    'tax_district_id': levy_cd,
                    'levy_cd': levy_cd,
                    'levy_cd_linked': levy_cd_linked,
                    'levy_rate': rate,
                    'levy_amount': levy,
                    'assessed_value': value,
                    'year': year,
                    'source': 'xls'
                }
                records.append(record)
        
        except Exception as e:
            logger.error(f"Error parsing XLS file {file_path}: {str(e)}")
            raise
        
        logger.info(f"Parsed {len(records)} records from XLS file")
        return LevyExportData(records, {'format': 'xls', 'year': year})
    
    @classmethod
    def _parse_xlsx(cls, file_path: Path) -> LevyExportData:
        """
        Parse an Excel .xlsx format levy export file.
        
        Args:
            file_path: Path to the XLSX file
            
        Returns:
            LevyExportData object containing the parsed data
        """
        logger.info(f"Parsing XLSX levy export file: {file_path}")
        
        try:
            # Open the workbook and select the first worksheet
            wb = openpyxl.load_workbook(file_path, data_only=True)
            ws = wb.active
            
            # Attempt to find header row
            header_row = None
            year = datetime.now().year
            
            for row_idx in range(1, min(21, ws.max_row + 1)):  # Check first 20 rows (1-indexed)
                row_values = [str(cell.value or '').strip().upper() for cell in ws[row_idx]]
                row_text = ' '.join(row_values)
                
                # Look for year in header rows
                if re.search(r'(LEVY|TAX)\s+YEAR\s*[:=\s]\s*(\d{4})', row_text, re.IGNORECASE):
                    year_match = re.search(r'(\d{4})', row_text)
                    if year_match:
                        year = int(year_match.group(1))
                        logger.info(f"Found year in header: {year}")
                
                # Look for column headers
                if 'LEVY CODE' in row_text and 'RATE' in row_text:
                    header_row = row_idx
                    logger.info(f"Found header row at index {header_row}")
                    break
            
            # If header not found, try to infer it
            if header_row is None:
                logger.warning("No header row found, attempting to infer column structure")
                header_row = 1
            
            # Get column indices
            header_values = [str(cell.value or '').strip().upper() for cell in ws[header_row]]
            
            col_indices = {
                'levy_cd': next((i for i, h in enumerate(header_values) if 'LEVY CODE' in h or 'CODE' in h), 0),
                'rate': next((i for i, h in enumerate(header_values) if 'RATE' in h), 1),
                'levy': next((i for i, h in enumerate(header_values) if 'LEVY' in h and 'CODE' not in h), 2),
                'value': next((i for i, h in enumerate(header_values) if 'VALUE' in h or 'ASSESSED' in h), 3)
            }
            
            # Parse data rows
            records = []
            for row_idx in range(header_row + 1, ws.max_row + 1):
                row_values = [cell.value for cell in ws[row_idx]]
                
                if not row_values or not row_values[col_indices['levy_cd']]:  # Skip empty rows
                    continue
                
                # Get values from appropriate columns
                levy_cd = str(row_values[col_indices['levy_cd']]).strip()
                
                # Handle potential formatting issues
                try:
                    rate = float(row_values[col_indices['rate']]) if row_values[col_indices['rate']] is not None else None
                except (ValueError, TypeError):
                    rate = None
                
                try:
                    levy = float(row_values[col_indices['levy']]) if row_values[col_indices['levy']] is not None else None
                except (ValueError, TypeError):
                    levy = None
                
                try:
                    value = float(row_values[col_indices['value']]) if row_values[col_indices['value']] is not None else None
                except (ValueError, TypeError):
                    value = None
                
                # Check for linked levy code
                levy_cd_linked = None
                if '/' in levy_cd:
                    levy_cd, levy_cd_linked = levy_cd.split('/', 1)
                
                record = {
                    'tax_district_id': levy_cd,
                    'levy_cd': levy_cd,
                    'levy_cd_linked': levy_cd_linked,
                    'levy_rate': rate,
                    'levy_amount': levy,
                    'assessed_value': value,
                    'year': year,
                    'source': 'xlsx'
                }
                records.append(record)
        
        except Exception as e:
            logger.error(f"Error parsing XLSX file {file_path}: {str(e)}")
            raise
        
        logger.info(f"Parsed {len(records)} records from XLSX file")
        return LevyExportData(records, {'format': 'xlsx', 'year': year})
    
    @classmethod
    def _parse_xml(cls, file_path: Path) -> LevyExportData:
        """
        Parse an XML format levy export file.
        
        Args:
            file_path: Path to the XML file
            
        Returns:
            LevyExportData object containing the parsed data
        """
        logger.info(f"Parsing XML levy export file: {file_path}")
        
        try:
            import xml.etree.ElementTree as ET
            
            tree = ET.parse(file_path)
            root = tree.getroot()
            
            # Extract year from metadata if available
            year = datetime.now().year
            year_elements = root.findall('.//year') or root.findall('.//Year')
            if year_elements:
                try:
                    year = int(year_elements[0].text)
                    logger.info(f"Found year in XML: {year}")
                except (ValueError, TypeError):
                    pass
            
            # Parse levy records
            records = []
            
            # Try different possible structures
            levy_elements = (
                root.findall('.//levy') or 
                root.findall('.//Levy') or 
                root.findall('.//levy_record') or
                root.findall('.//LevyRecord')
            )
            
            if not levy_elements:
                logger.warning("No levy records found in XML structure")
            
            for levy_elem in levy_elements:
                # Extract code
                code_elem = (
                    levy_elem.find('./code') or
                    levy_elem.find('./Code') or
                    levy_elem.find('./levy_code') or
                    levy_elem.find('./LevyCode')
                )
                
                if code_elem is None or not code_elem.text:
                    continue
                
                levy_cd = code_elem.text.strip()
                
                # Extract rate
                rate_elem = (
                    levy_elem.find('./rate') or
                    levy_elem.find('./Rate')
                )
                rate = None
                if rate_elem is not None and rate_elem.text:
                    try:
                        rate = float(rate_elem.text)
                    except (ValueError, TypeError):
                        pass
                
                # Extract levy
                levy_amount_elem = (
                    levy_elem.find('./amount') or
                    levy_elem.find('./Amount') or
                    levy_elem.find('./levy_amount') or
                    levy_elem.find('./LevyAmount')
                )
                levy_amount = None
                if levy_amount_elem is not None and levy_amount_elem.text:
                    try:
                        levy_amount = float(levy_amount_elem.text)
                    except (ValueError, TypeError):
                        pass
                
                # Extract value
                value_elem = (
                    levy_elem.find('./value') or
                    levy_elem.find('./Value') or
                    levy_elem.find('./assessed_value') or
                    levy_elem.find('./AssessedValue')
                )
                value = None
                if value_elem is not None and value_elem.text:
                    try:
                        value = float(value_elem.text)
                    except (ValueError, TypeError):
                        pass
                
                # Check for linked levy code
                levy_cd_linked = None
                if '/' in levy_cd:
                    levy_cd, levy_cd_linked = levy_cd.split('/', 1)
                
                record = {
                    'tax_district_id': levy_cd,
                    'levy_cd': levy_cd,
                    'levy_cd_linked': levy_cd_linked,
                    'levy_rate': rate,
                    'levy_amount': levy_amount,
                    'assessed_value': value,
                    'year': year,
                    'source': 'xml'
                }
                records.append(record)
        
        except Exception as e:
            logger.error(f"Error parsing XML file {file_path}: {str(e)}")
            raise
        
        logger.info(f"Parsed {len(records)} records from XML file")
        return LevyExportData(records, {'format': 'xml', 'year': year})